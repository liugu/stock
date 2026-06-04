#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
每日自动化任务脚本
功能：
1. 数据更新：调用 execute_daily_job.py 更新数据
2. 策略选股：运行策略选股并生成结果Excel
3. Server酱推送：将选股结果发送到Server酱

使用方法：
    python daily_task.py                    # 执行所有任务
    python daily_task.py --no-update        # 跳过数据更新
    python daily_task.py --no-selection     # 跳过策略选股
    python daily_task.py --no-push          # 跳过推送
    python daily_task.py --date 2024-01-15  # 指定日期
"""

import argparse
import datetime
import logging
import os
import sys
import time
import json
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 添加项目路径
cpath_current = os.path.dirname(os.path.dirname(__file__))
cpath = os.path.abspath(os.path.join(cpath_current, os.pardir))
sys.path.append(cpath)

import instock.lib.database as mdb
import instock.core.tablestructure as tbs
import instock.lib.trade_time as trd

__author__ = 'liugu'
__date__ = '2024/05/10'

# 配置文件路径
CONFIG_FILE = os.path.join(os.path.dirname(__file__), 'daily_task_config.json')

# 日志配置
LOG_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'log')
if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)

logging.basicConfig(
    format='%(asctime)s %(levelname)s %(message)s',
    filename=os.path.join(LOG_DIR, 'daily_task.log'),
    level=logging.INFO
)
logger = logging.getLogger(__name__)


def load_config():
    """加载配置文件"""
    default_config = {
        "serverchan_key": "SCT347969TWcw4Zztqp8nMDiwYE4ik2EOW",
        "output_dir": "./output",
        "strategies": [
            {"name": "cn_stock_strategy_enter", "cn": "放量上涨"},
            {"name": "cn_stock_strategy_keep_increasing", "cn": "均线多头"},
            {"name": "cn_stock_strategy_parking_apron", "cn": "停机坪"},
            {"name": "cn_stock_strategy_backtrace_ma250", "cn": "回踩年线"},
            {"name": "cn_stock_strategy_breakthrough_platform", "cn": "突破平台"},
            {"name": "cn_stock_strategy_low_backtrace_increase", "cn": "无大幅回撤"},
            {"name": "cn_stock_strategy_turtle_trade", "cn": "海龟交易法则"},
            {"name": "cn_stock_strategy_high_tight_flag", "cn": "高而窄的旗形"},
            {"name": "cn_stock_strategy_climax_limitdown", "cn": "放量跌停"},
            {"name": "cn_stock_strategy_low_atr", "cn": "低ATR成长"},
            {"name": "cn_stock_strategy_consecutive_small_bullish", "cn": "连续小阳线"}
        ]
    }

    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)
                # 合并配置
                for key in default_config:
                    if key not in config:
                        config[key] = default_config[key]
                return config
        except Exception as e:
            logger.error(f"加载配置文件失败: {e}")
            return default_config
    return default_config


def run_data_update(date=None):
    """执行数据更新"""
    logger.info("=" * 50)
    logger.info("开始执行数据更新任务")
    start_time = time.time()

    try:
        # 导入并执行每日作业
        import instock.job.execute_daily_job as daily_job

        if date:
            # 指定日期
            logger.info(f"更新指定日期数据: {date}")
            year, month, day = date.split('-')
            run_date = datetime.date(int(year), int(month), int(day))

            # 直接调用各模块
            import instock.job.init_job as bj
            import instock.job.basic_data_daily_job as hdj
            import instock.job.basic_data_other_daily_job as hdtj
            import instock.job.basic_data_after_close_daily_job as acdj
            import instock.job.indicators_data_daily_job as gdj
            import instock.job.strategy_data_daily_job as sdj
            import instock.job.backtest_data_daily_job as bdj
            import instock.job.klinepattern_data_daily_job as kdj
            import instock.job.selection_data_daily_job as sddj
            import concurrent.futures

            bj.main()
            hdj.main()
            sddj.main()

            with concurrent.futures.ThreadPoolExecutor() as executor:
                executor.submit(hdtj.main)
                executor.submit(gdj.main)
                executor.submit(kdj.main)
                executor.submit(sdj.main)

            bdj.main()
            acdj.main()
        else:
            # 使用默认逻辑
            daily_job.main()

        elapsed = time.time() - start_time
        logger.info(f"数据更新完成，耗时: {elapsed:.2f} 秒")
        return True

    except Exception as e:
        logger.error(f"数据更新失败: {e}")
        return False


def get_strategy_results(date, config):
    """获取策略选股结果"""
    logger.info("=" * 50)
    logger.info("开始获取策略选股结果")

    results = {}

    try:
        conn = mdb.get_connection()
        if conn is None:
            logger.error("数据库连接失败")
            return results

        with conn.cursor() as cursor:
            for strategy in config.get('strategies', []):
                table_name = strategy['name']
                strategy_cn = strategy['cn']

                try:
                    # 查询策略结果
                    sql = f"""
                        SELECT s.date, s.code, s.name,
                               sp.new_price, sp.change_rate, sp.turnoverrate,
                               sp.total_market_cap, sp.industry
                        FROM `{table_name}` s
                        LEFT JOIN cn_stock_spot sp ON s.code = sp.code AND s.date = sp.date
                        WHERE s.date = '{date}'
                        ORDER BY sp.change_rate DESC
                    """
                    cursor.execute(sql)
                    rows = cursor.fetchall()

                    if rows:
                        df = pd.DataFrame(rows, columns=[
                            'date', 'code', 'name', 'new_price',
                            'change_rate', 'turnoverrate',
                            'total_market_cap', 'industry'
                        ])
                        results[strategy_cn] = df
                        logger.info(f"策略 [{strategy_cn}] 选出 {len(df)} 只股票")
                    else:
                        logger.info(f"策略 [{strategy_cn}] 无选股结果")

                except Exception as e:
                    logger.error(f"查询策略 {strategy_cn} 失败: {e}")
                    continue

    except Exception as e:
        logger.error(f"获取策略选股结果失败: {e}")

    return results


def generate_excel_report(date, results, config):
    """生成Excel报告"""
    logger.info("=" * 50)
    logger.info("开始生成Excel报告")

    output_dir = config.get('output_dir', './output')
    if not os.path.isabs(output_dir):
        output_dir = os.path.join(os.path.dirname(__file__), output_dir)

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    filename = os.path.join(output_dir, f"选股报告_{date}.xlsx")

    try:
        wb = Workbook()
        # 删除默认sheet
        wb.remove(wb.active)

        # 样式定义
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font_white = Font(bold=True, size=11, color='FFFFFF')
        red_font = Font(color='FF0000')
        green_font = Font(color='00B050')
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 汇总sheet
        summary_ws = wb.create_sheet('汇总')
        summary_ws.append(['策略名称', '选股数量'])
        for cell in summary_ws[1]:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        total_count = 0
        row_idx = 2
        for strategy_name, df in results.items():
            count = len(df)
            total_count += count
            summary_ws.append([strategy_name, count])
            for cell in summary_ws[row_idx]:
                cell.alignment = center_align
                cell.border = thin_border
            row_idx += 1

        summary_ws.append(['总计', total_count])
        for cell in summary_ws[row_idx]:
            cell.font = Font(bold=True)
            cell.alignment = center_align
            cell.border = thin_border

        # 调整列宽
        summary_ws.column_dimensions['A'].width = 20
        summary_ws.column_dimensions['B'].width = 12

        # 各策略详情sheet
        for strategy_name, df in results.items():
            if df.empty:
                continue

            # sheet名称最长31字符
            sheet_name = strategy_name[:31].replace('/', '_').replace('\\', '_')
            ws = wb.create_sheet(sheet_name)

            # 写入表头
            headers = ['代码', '名称', '最新价', '涨跌幅(%)', '换手率(%)', '市值(亿)', '行业']
            ws.append(headers)

            for col_idx, cell in enumerate(ws[1], 1):
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # 写入数据
            for _, row in df.iterrows():
                row_data = [
                    row['code'],
                    row['name'],
                    round(row['new_price'], 2) if pd.notna(row['new_price']) else '',
                    round(row['change_rate'], 2) if pd.notna(row['change_rate']) else '',
                    round(row['turnoverrate'], 2) if pd.notna(row['turnoverrate']) else '',
                    round(row['total_market_cap'] / 100000000, 2) if pd.notna(row['total_market_cap']) else '',
                    row['industry'] if pd.notna(row['industry']) else ''
                ]
                ws.append(row_data)

                # 设置样式
                for col_idx, cell in enumerate(ws[ws.max_row], 1):
                    cell.alignment = center_align
                    cell.border = thin_border

                    # 涨跌幅列着色
                    if col_idx == 4 and row_data[3]:
                        if row_data[3] > 0:
                            cell.font = red_font
                        elif row_data[3] < 0:
                            cell.font = green_font

            # 调整列宽
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 15

        wb.save(filename)
        logger.info(f"Excel报告已生成: {filename}")
        return filename

    except Exception as e:
        logger.error(f"生成Excel报告失败: {e}")
        return None


def send_serverchan_message(date, results, config):
    """发送Server酱消息"""
    serverchan_key = config.get('serverchan_key', '')

    if not serverchan_key:
        logger.warning("未配置Server酱key，跳过推送")
        return False

    logger.info("=" * 50)
    logger.info("开始发送Server酱消息")

    try:
        # 构建消息内容
        total_count = sum(len(df) for df in results.values())

        if total_count == 0:
            content = f"## 📊 选股报告 - {date}\n\n今日策略选股结果为空，无符合条件的股票。"
        else:
            content = f"## 📊 选股报告 - {date}\n\n"
            content += f"**策略选股汇总：**\n\n"

            for strategy_name, df in results.items():
                count = len(df)
                if count > 0:
                    content += f"- **{strategy_name}**: {count} 只\n"

            content += f"\n**总计**: {total_count} 只\n\n"

            # 添加涨幅前5的股票
            all_stocks = []
            for strategy_name, df in results.items():
                for _, row in df.iterrows():
                    all_stocks.append({
                        'strategy': strategy_name,
                        'code': row['code'],
                        'name': row['name'],
                        'change_rate': row['change_rate'] if pd.notna(row['change_rate']) else 0
                    })

            if all_stocks:
                # 按涨幅排序
                all_stocks.sort(key=lambda x: x['change_rate'], reverse=True)
                top_stocks = all_stocks[:5]

                content += "**涨幅前5：**\n\n"
                for i, stock in enumerate(top_stocks, 1):
                    change_str = f"+{stock['change_rate']:.2f}%" if stock['change_rate'] > 0 else f"{stock['change_rate']:.2f}%"
                    content += f"{i}. {stock['name']}({stock['code']}): {change_str}\n"

        # 发送消息
        url = f"https://sctapi.ftqq.com/{serverchan_key}.send"
        data = {"title": f"选股报告 - {date}", "desp": content}
        response = requests.post(url, json=data, timeout=10)

        result = response.json()
        if result.get('code') == 0:
            logger.info("Server酱消息发送成功")
            return True
        else:
            logger.error(f"Server酱消息发送失败: {result.get('message')}")
            return False
    except Exception as e:
        logger.error(f"发送Server酱消息失败: {e}")
        return False


def main():
    """主函数"""
    parser = argparse.ArgumentParser(description='每日自动化任务脚本')
    parser.add_argument('--no-update', action='store_true', help='跳过数据更新')
    parser.add_argument('--no-selection', action='store_true', help='跳过策略选股')
    parser.add_argument('--no-push', action='store_true', help='跳过推送')
    parser.add_argument('--date', type=str, help='指定日期 (格式: YYYY-MM-DD)')
    parser.add_argument('--config', type=str, help='指定配置文件路径')

    args = parser.parse_args()

    # 加载配置
    global CONFIG_FILE
    if args.config:
        CONFIG_FILE = args.config
    config = load_config()

    # 确定日期
    if args.date:
        target_date = args.date
    else:
        # 获取最近交易日
        run_date, _ = trd.get_trade_date_last()
        target_date = run_date.strftime('%Y-%m-%d')

    logger.info("=" * 60)
    logger.info(f"每日自动化任务开始 - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"目标日期: {target_date}")
    logger.info("=" * 60)

    start_time = time.time()
    success = True

    # 1. 数据更新
    if not args.no_update:
        if not run_data_update(args.date):
            success = False
    else:
        logger.info("跳过数据更新")

    # 2. 策略选股
    results = {}
    if not args.no_selection:
        results = get_strategy_results(target_date, config)
        if results:
            # 生成Excel报告
            excel_file = generate_excel_report(target_date, results, config)
        else:
            logger.warning("未获取到策略选股结果")
    else:
        logger.info("跳过策略选股")

    # 3. Server酱推送
    if not args.no_push:
        send_serverchan_message(target_date, results, config)
    else:
        logger.info("跳过推送")

    elapsed = time.time() - start_time
    logger.info("=" * 60)
    logger.info(f"每日自动化任务完成，总耗时: {elapsed:.2f} 秒")
    logger.info("=" * 60)

    return 0 if success else 1


if __name__ == '__main__':
    sys.exit(main())
